# -*- coding: utf-8 -*-
"""Generate detailed backlog (subtasks), CSVs, multi-sheet Excel, LOC template, and backlog.md."""
import csv
import importlib.util
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
CSV_PATH = ROOT / "backlog-full.csv"
EPICS_CSV_PATH = ROOT / "backlog-epics.csv"
DETAILED_XLSX_PATH = ROOT / "backlog-orcaxcare.xlsx"
TEMPLATE_CSV_PATH = ROOT / "template-loc-orcaxcare-filled.csv"
TEMPLATE_XLSX_PATH = ROOT / "template-loc-orcaxcare-filled.xlsx"
MD_PATH = ROOT / "backlog.md"
SPRINT_CSV_PATHS = {n: ROOT / f"backlog-sprint-{n}.csv" for n in range(1, 5)}

_spec = importlib.util.spec_from_file_location("backlog_subtasks", ROOT / "backlog_subtasks.py")
_subm = importlib.util.module_from_spec(_spec)
assert _spec.loader is not None
_spec.loader.exec_module(_subm)
SUBTASKS_BY_PARENT: dict[int, list[tuple[str, str, int, str, str]]] = _subm.SUBTASKS_BY_PARENT

# Aligns with instructor template: https://docs.google.com/spreadsheets/d/1bf83uwr-moBSOkmPZdaZ8URypso6dh0F/edit?gid=1114776703
ITERATION_NAME = "OrcaXCare (WDP301)"
TEMPLATE_HEADERS = [
    "#",
    "Parent",
    "Screen / Function",
    "Feature",
    "Screen/Function Description",
    "In Charge",
    "Status",
    "Sprint",
    "SRS",
    "SDS",
    "LOC",
    "Code Demo Comments",
    "Complexity",
    "Quality",
]
NCOLS = len(TEMPLATE_HEADERS)

# Epic row: parent_id, Screen_Function, epic_feature, epic_description, Owner, Status, SRS, SDS, total_LOC, demo, epic_complexity, Quality, Sprint
EPICS = [
    (1, "Queue screen (patient + doctor)", "Real-time queue", "Doctor clicks Next; Socket.io syncs the queue number to the patient screen immediately.", "TruongNTCE180140", "To Do", "SRS §", "SDS §", 180, "Demo two browsers in the same session.", "High", "", 4),
    (2, "Admin / Doctor — Schedule", "Appointment slot scheduling", "Algorithm generates time slots from doctor shifts; prevents duplicate internal slots.", "TruongNTCE180140", "To Do", "SRS §", "SDS §", 200, "Show slot grid by day.", "High", "", 3),
    (3, "Server — API architecture", "Architecture & auth middleware", "Express boilerplate; layered JWT middleware; centralized error handling.", "TruongNTCE180140", "To Do", "SRS §", "SDS §", 120, "Sample Postman collection.", "Medium", "", 1),
    (4, "Portal + API", "Registration & login", "Accounts; password hashing; safe token issue and revocation.", "TruongNTCE180140", "To Do", "SRS §", "SDS §", 150, "Short video of registration flow.", "Medium", "", 2),
    (5, "Portal — Find doctor", "Smart search", "Search by name; specialty; symptom hints (regex / index).", "TruongNTCE180140", "To Do", "SRS §", "SDS §", 100, "A few sample queries.", "Medium", "", 3),
    (6, "Portal — Profile", "Personal profile management", "Update info; avatar; basic medical history.", "TruongNTCE180140", "To Do", "SRS §", "SDS §", 90, "Validated form.", "Standard", "", 2),
    (7, "Payments", "Online payment", "MoMo or VNPay sandbox integration; callback handling.", "TanhNTCE182341", "To Do", "SRS §", "SDS §", 220, "Successful test transaction.", "High", "", 4),
    (8, "Wallet + API", "Internal medical wallet", "Balance; atomic credit/debit; non-negative balance.", "TanhNTCE182341", "To Do", "SRS §", "SDS §", 180, "Unit tests for wallet flow.", "High", "", 4),
    (9, "Portal — Health insurance", "Health insurance management", "Store insurance card; optional OCR; apply discount per rules.", "TanhNTCE182341", "To Do", "SRS §", "SDS §", 160, "Cases with and without insurance.", "Medium", "", 4),
    (10, "Portal — History", "Transaction history", "List of top-ups and spends; filter by date and type.", "TanhNTCE182341", "To Do", "SRS §", "SDS §", 110, "Optional CSV export.", "Medium", "", 4),
    (11, "Booking + Wallet", "Automatic refund", "Cancel before deadline; refund to wallet per policy.", "TanhNTCE182341", "To Do", "SRS §", "SDS §", 130, "Two cancel cases: inside and outside policy window.", "Medium", "", 4),
    (12, "Server — RBAC", "RBAC authorization", "Medical record view policy; stock approval; consistent route guards.", "TanhNTCE182341", "To Do", "SRS §", "SDS §", 100, "Permission matrix.", "Standard", "", 2),
    (13, "Doctor — EMR", "EMR medical record", "Diagnosis timeline; quick lookup by patient.", "ThangNDCE180608", "To Do", "SRS §", "SDS §", 240, "Three sample encounters.", "High", "", 4),
    (14, "Admin — Stock", "Pharmacy stock management", "Stock in/out; low-stock and expiry warnings.", "ThangNDCE180608", "To Do", "SRS §", "SDS §", 200, "Stock table view.", "High", "", 4),
    (15, "Admin — Staff", "Doctor management", "CRUD medical staff; departments; specialties.", "ThangNDCE180608", "To Do", "SRS §", "SDS §", 160, "CRUD demo.", "Medium", "", 3),
    (16, "Doctor — Imaging", "Imaging library", "Cloudinary upload for X-ray / MRI; album per visit.", "ThangNDCE180608", "To Do", "SRS §", "SDS §", 140, "Upload one image.", "Medium", "", 4),
    (17, "Admin — Facilities", "Clinic room management", "Functional rooms; equipment per room.", "ThangNDCE180608", "To Do", "SRS §", "SDS §", 120, "Room layout diagram.", "Standard", "", 1),
    (18, "Admin — Catalog", "ICD catalog", "Standard disease codes; search when entering diagnosis.", "ThangNDCE180608", "To Do", "SRS §", "SDS §", 100, "Sample CSV import.", "Standard", "", 3),
    (19, "Portal — Booking", "Booking flow", "Four steps: Department → Doctor → Time → Payment.", "ThangDQCE182036", "To Do", "SRS §", "SDS §", 220, "Screen recording.", "High", "", 3),
    (20, "API Booking", "Double-booking check", "Prevent booking a slot already taken.", "ThangDQCE182036", "To Do", "SRS §", "SDS §", 120, "Concurrent test with two users.", "High", "", 3),
    (21, "Portal — Home", "Patient home page", "Featured services; featured doctors; health news.", "ThangDQCE182036", "To Do", "SRS §", "SDS §", 130, "Responsive landing.", "Medium", "", 3),
    (22, "Notifications", "Web Push", "Appointment reminders; lab results available.", "ThangDQCE182036", "To Do", "SRS §", "SDS §", 140, "Subscribe in Chrome.", "Medium", "", 4),
    (23, "Portal — Review", "Doctor rating", "After visit; rating and comment.", "ThangDQCE182036", "To Do", "SRS §", "SDS §", 100, "One sample review.", "Medium", "", 4),
    (24, "Portal — Favorites", "Preferred doctor list", "Shortcuts for faster re-booking.", "ThangDQCE182036", "To Do", "SRS §", "SDS §", 80, "Add/remove favorite.", "Standard", "", 3),
    (25, "Admin — Reports", "Revenue report", "MongoDB aggregation + Chart.js charts.", "KhoaNNCE181612", "To Do", "SRS §", "SDS §", 180, "Monthly dashboard.", "High", "", 4),
    (26, "Doctor — Prescriptions", "Prescription PDF export", "Medical-standard template from DB data.", "KhoaNNCE181612", "To Do", "SRS §", "SDS §", 160, "Sample PDF file.", "High", "", 4),
    (27, "Portal + Kiosk", "QR verification", "QR on prescription and record for scan verification.", "KhoaNNCE181612", "To Do", "SRS §", "SDS §", 100, "Scan with phone camera.", "Medium", "", 4),
    (28, "System — Email", "Automated email", "SMTP for visit results, prescription, follow-up.", "KhoaNNCE181612", "To Do", "SRS §", "SDS §", 120, "MailHog or log output.", "Medium", "", 4),
    (29, "Portal — Maps", "Branch map", "Maps API directions to clinic.", "KhoaNNCE181612", "To Do", "SRS §", "SDS §", 100, "One map marker.", "Medium", "", 3),
    (30, "Admin + Portal", "Complaint handling", "Intake; processing status.", "KhoaNNCE181612", "To Do", "SRS §", "SDS §", 110, "Ticket workflow.", "Standard", "", 4),
    (31, "Server — Quality", "Input validation", "Joi or Zod on main APIs.", "Team", "To Do", "SRS §", "SDS §", 150, "Example 400 error response.", "Medium", "", 1),
    (32, "Server — Quality", "Unified error handling", "Error codes; messages; stack traces logged dev-only.", "Team", "To Do", "SRS §", "SDS §", 80, "Postman error response shape.", "Medium", "", 1),
    (33, "DevOps", "Seed & light migration", "Script to seed sample ICD codes, users, and clinics.", "Team", "To Do", "SRS §", "SDS §", 60, "npm run seed.", "Standard", "", 1),
    (34, "Docs", "Short API contract", "OpenAPI or endpoint table for the team.", "Team", "To Do", "SRS §", "SDS §", 40, "Link in README.", "Standard", "", 1),
    (35, "QA", "Supertest smoke", "Minimal path: health, auth, booking, wallet.", "Team", "To Do", "SRS §", "SDS §", 100, "CI runs tests.", "Medium", "", 4),
]


def expand_rows() -> list[tuple]:
    """Flatten EPICS + SUBTASKS_BY_PARENT into subtask rows."""
    missing = [e[0] for e in EPICS if e[0] not in SUBTASKS_BY_PARENT]
    if missing:
        raise ValueError(f"SUBTASKS_BY_PARENT missing keys: {missing}")
    out: list[tuple] = []
    n = 0
    for epic in EPICS:
        pid, sf, epic_feat, epic_desc, owner, status, srs, sds, _total_loc, _demo, _comp, qual, sprint = epic
        for feature, desc, loc, demo, comp in SUBTASKS_BY_PARENT[pid]:
            n += 1
            out.append((n, pid, sf, feature, desc, owner, status, sprint, srs, sds, loc, demo, qual, comp))
    return out


ROWS = expand_rows()


def pad(row: list) -> list:
    r = list(row)
    while len(r) < NCOLS:
        r.append("")
    return r[:NCOLS]


def demo_label(complexity: str) -> str:
    if complexity == "High":
        return "Complex"
    if complexity == "Medium":
        return "Medium"
    return "Simple"


def row_csv_values(row: tuple) -> list:
    """Order for backlog-full.csv and backlog-sprint-*.csv (subtasks)."""
    sub_id, parent_id, sf, feat, desc, owner, status, sprint, srs, sds, loc, demo, qual, comp = row
    return [sub_id, parent_id, sf, feat, desc, owner, status, sprint, srs, sds, loc, demo, comp, qual]


def write_csv() -> None:
    headers = [
        "#",
        "Parent_ID",
        "Screen_Function",
        "Feature",
        "Description",
        "In_Charge",
        "Status",
        "Sprint",
        "SRS",
        "SDS",
        "LOC",
        "Code_Demo_Comments",
        "Complexity",
        "Quality",
    ]
    with CSV_PATH.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for row in ROWS:
            w.writerow(row_csv_values(row))


def write_epics_csv() -> None:
    """Original 35 high-level backlog rows (for slides / teacher macro view)."""
    headers = [
        "Parent_ID",
        "Screen_Function",
        "Epic_Feature",
        "Description",
        "In_Charge",
        "Status",
        "Sprint",
        "SRS",
        "SDS",
        "LOC",
        "Code_Demo_Comments",
        "Complexity",
        "Quality",
    ]
    with EPICS_CSV_PATH.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for epic in EPICS:
            pid, sf, feat, desc, owner, status, srs, sds, loc, demo, comp, qual, sprint = epic
            w.writerow([pid, sf, feat, desc, owner, status, sprint, srs, sds, loc, demo, comp, qual])


def write_sprint_csvs() -> None:
    headers = [
        "#",
        "Parent_ID",
        "Screen_Function",
        "Feature",
        "Description",
        "In_Charge",
        "Status",
        "Sprint",
        "SRS",
        "SDS",
        "LOC",
        "Code_Demo_Comments",
        "Complexity",
        "Quality",
    ]
    for n, path in SPRINT_CSV_PATHS.items():
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for row in ROWS:
                if row[7] == n:
                    w.writerow(row_csv_values(row))


def write_template_loc_csv() -> None:
    """CSV matching LOC Evaluation sheet: 3 instruction rows + header + data (columns A–N)."""
    row1 = pad(
        [
            f"LOC Evaluation for {ITERATION_NAME}",
        ]
    )
    row2 = pad(
        [
            "Columns A–K: # through LOC (subtasks). B = Parent epic id (1–35). H = Sprint (1–4).",
        ]
    )
    row3 = pad(
        [
            "Teacher fills columns L–N (Code Demo Comments, Complexity, Quality); pivot by In Charge for LOC totals.",
        ]
    )
    with TEMPLATE_CSV_PATH.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(row1)
        w.writerow(row2)
        w.writerow(row3)
        w.writerow(TEMPLATE_HEADERS)
        for row in ROWS:
            sub_id, parent_id, sf, feat, desc, owner, status, sprint, srs, sds, loc, demo, qual, comp = row
            lvl = demo_label(comp)
            code_demo = f"{lvl}: {demo}"
            w.writerow(
                pad(
                    [
                        sub_id,
                        parent_id,
                        sf,
                        feat,
                        desc,
                        owner,
                        status,
                        sprint,
                        srs,
                        sds,
                        loc,
                        code_demo,
                        comp,
                        qual,
                    ]
                )
            )


DETAILED_HEADERS = [
    "#",
    "Parent_ID",
    "Screen_Function",
    "Feature",
    "Description",
    "In_Charge",
    "Status",
    "Sprint",
    "SRS",
    "SDS",
    "LOC",
    "Code_Demo_Comments",
    "Complexity",
    "Quality",
]


def _style_subtask_sheet(ws, header_row: int, data_rows: list[tuple]) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    ncols = len(DETAILED_HEADERS)
    for col, title in enumerate(DETAILED_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    data_start = header_row + 1
    for offset, row in enumerate(data_rows):
        ridx = data_start + offset
        for col, val in enumerate(row_csv_values(row), start=1):
            cell = ws.cell(row=ridx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 11:
                cell.number_format = "0"
    wmap = {1: 6, 2: 6, 3: 28, 4: 30, 5: 48, 6: 14, 7: 10, 8: 8, 9: 8, 10: 8, 11: 8, 12: 38, 13: 12, 14: 8}
    for col, w in wmap.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    if data_rows:
        last_data = header_row + len(data_rows)
        ws.freeze_panes = f"A{data_start}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{last_data}"


def write_backlog_workbook() -> None:
    """Multi-sheet workbook: Epics (35), All subtasks, Sprint_1–4."""
    wb = Workbook()
    wb.remove(wb.active)

    ws_ep = wb.create_sheet("Epics", 0)
    epic_headers = [
        "Parent_ID",
        "Screen_Function",
        "Epic_Feature",
        "Description",
        "In_Charge",
        "Status",
        "Sprint",
        "SRS",
        "SDS",
        "LOC",
        "Code_Demo_Comments",
        "Complexity",
        "Quality",
        "Subtask_count",
        "Subtask_LOC_sum",
    ]
    fill = PatternFill("solid", fgColor="E2EFDA")
    for col, title in enumerate(epic_headers, start=1):
        c = ws_ep.cell(row=1, column=col, value=title)
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, epic in enumerate(EPICS, start=2):
        pid, sf, feat, desc, owner, status, srs, sds, loc, demo, comp, qual, sprint = epic
        parts = SUBTASKS_BY_PARENT[pid]
        vals = [pid, sf, feat, desc, owner, status, sprint, srs, sds, loc, demo, comp, qual, len(parts), sum(p[2] for p in parts)]
        for col, val in enumerate(vals, start=1):
            ws_ep.cell(row=i, column=col, value=val).alignment = Alignment(wrap_text=True, vertical="top")
    for col in range(1, 16):
        ws_ep.column_dimensions[get_column_letter(col)].width = 14 if col != 4 else 40
    ws_ep.freeze_panes = "A2"
    ws_ep.auto_filter.ref = f"A1:{get_column_letter(len(epic_headers))}{1 + len(EPICS)}"

    ws_all = wb.create_sheet("All")
    _style_subtask_sheet(ws_all, 1, ROWS)

    for n in range(1, 5):
        ws = wb.create_sheet(f"Sprint_{n}")
        _style_subtask_sheet(ws, 1, [r for r in ROWS if r[7] == n])

    wb.save(DETAILED_XLSX_PATH)


def write_template_loc_xlsx() -> None:
    """Excel (.xlsx) with the same LOC template content for viewing or submission."""
    wb = Workbook()
    ws = wb.active
    ws.title = "LOC_OrcaXCare"

    last_letter = get_column_letter(NCOLS)
    instructions = [
        f"LOC Evaluation for {ITERATION_NAME}",
        "Columns A–K: # through LOC (subtasks). B = Parent epic id (1–35). H = Sprint (1–4).",
        "Teacher fills columns L–N (Code Demo Comments, Complexity, Quality); pivot by In Charge for LOC totals.",
    ]
    for i, text in enumerate(instructions, start=1):
        ws.merge_cells(f"A{i}:{last_letter}{i}")
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = Font(bold=(i == 1), size=12 if i == 1 else 10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    header_row = 4
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)
    for col, title in enumerate(TEMPLATE_HEADERS, start=1):
        c = ws.cell(row=header_row, column=col, value=title)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")

    data_start = header_row + 1
    for offset, row in enumerate(ROWS):
        ridx = data_start + offset
        sub_id, parent_id, sf, feat, desc, owner, status, sprint, srs, sds, loc, demo, qual, comp = row
        lvl = demo_label(comp)
        code_demo = f"{lvl}: {demo}"
        values = [sub_id, parent_id, sf, feat, desc, owner, status, sprint, srs, sds, loc, code_demo, comp, qual]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=ridx, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 11:
                cell.number_format = "0"

    widths = {1: 6, 2: 6, 3: 28, 4: 24, 5: 48, 6: 16, 7: 10, 8: 8, 9: 10, 10: 10, 11: 8, 12: 40, 13: 12, 14: 10}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    last_data_row = header_row + len(ROWS)
    ws.freeze_panes = f"A{data_start}"
    ws.auto_filter.ref = f"A{header_row}:{last_letter}{last_data_row}"

    wb.save(TEMPLATE_XLSX_PATH)


def write_md() -> None:
    n_sub = len(ROWS)
    counts = Counter(r[7] for r in ROWS)
    intro = f"""# OrcaXCare — Backlog

**{n_sub} subtasks** split from **35 epics** (course scope + shared technical items). Sprint **1–4** follows [`sprint-plan.md`](./sprint-plan.md).

| Artifact | Purpose |
|----------|---------|
| [`backlog-orcaxcare.xlsx`](./backlog-orcaxcare.xlsx) | **Excel:** sheets *Epics*, *All*, *Sprint_1* … *Sprint_4* |
| [`backlog-full.csv`](./backlog-full.csv) | All subtasks (import to Sheets) |
| [`backlog-epics.csv`](./backlog-epics.csv) | 35 epic rows only (slides / macro) |
| [`backlog-sprint-1.csv`](./backlog-sprint-1.csv) … [`backlog-sprint-4.csv`](./backlog-sprint-4.csv) | Subtasks filtered by sprint |
| [`backlog_subtasks.py`](./backlog_subtasks.py) | Edit subtask text / LOC splits |

Regenerate: `python docs/generate_backlog.py`

---

## Epics (35 rows)

| Parent | Screen / Function | Epic feature | Sprint | Owner | LOC (epic) | #Sub |
|--------|--------------------|--------------|--------|-------|------------|------|
"""
    lines = [intro]
    for epic in EPICS:
        pid, sf, feat, _desc, owner, _status, _srs, _sds, loc, _demo, _comp, _qual, sprint = epic
        n_ch = len(SUBTASKS_BY_PARENT[pid])
        sf_esc = sf.replace("|", "/")
        feat_esc = feat.replace("|", "/")
        lines.append(f"| {pid} | {sf_esc} | {feat_esc} | {sprint} | {owner} | {loc} | {n_ch} |\n")
    lines.append("\n---\n\n## Sprint subtask counts\n\n")
    for n in range(1, 5):
        lines.append(f"- **Sprint {n}:** {counts[n]} subtasks ([`backlog-sprint-{n}.csv`](./backlog-sprint-{n}.csv))\n")
    lines.append(
        "\n---\n\n## Notes\n\n"
        "- **LOC (subtask):** rough slice of the epic LOC; adjust in `backlog_subtasks.py`.\n"
        "- **Sprint reassignment:** change the last number on each tuple in `EPICS` inside `generate_backlog.py`, then rerun.\n"
    )
    MD_PATH.write_text("".join(lines), encoding="utf-8")


def main() -> None:
    write_csv()
    write_epics_csv()
    write_sprint_csvs()
    write_template_loc_csv()
    write_md()
    sprint_paths = ", ".join(str(p) for p in SPRINT_CSV_PATHS.values())
    warnings: list[str] = []
    wb_ok = False
    template_ok = False

    try:
        write_backlog_workbook()
        wb_ok = True
    except PermissionError:
        warnings.append(
            f"Could not write {DETAILED_XLSX_PATH} (file may be open). Close Excel and run again."
        )

    try:
        write_template_loc_xlsx()
        template_ok = True
    except PermissionError:
        warnings.append(
            f"Could not write {TEMPLATE_XLSX_PATH} (file may be open). Close Excel and run again."
        )

    print(f"Wrote {CSV_PATH}, {EPICS_CSV_PATH}, {sprint_paths}, {TEMPLATE_CSV_PATH}, {MD_PATH}")
    if wb_ok:
        print(f"Wrote {DETAILED_XLSX_PATH}")
    if template_ok:
        print(f"Wrote {TEMPLATE_XLSX_PATH}")
    for w in warnings:
        print(f"WARNING: {w}")


if __name__ == "__main__":
    main()
